import numpy as np
import pandas as pd
import scipy.optimize as opt
from openpyxl import load_workbook
import os
import sys

def linear_ev_market_share_smoothed(target_fleet_ev_percentage, start_year, target_year, start_ev_percentage, age_distribution):
    """
    Computes EV market share growth with a smooth start and natural transition.

    Parameters:
    - target_fleet_ev_percentage: The desired fleet EV share in the target year.
    - start_year: The year from which EV adoption starts.
    - target_year: The year to reach the target fleet EV share.
    - start_ev_percentage: The initial EV market share in the start year.
    - age_distribution: The extracted age fraction data from the input file.
    
    Returns:
    - DataFrame with yearly EV market share and cumulative fleet EV percentage.
    """
    years = list(range(start_year, target_year + 1))

    def compute_required_growth(alpha, beta, transition_start, transition_end):
        fleet_composition = np.zeros(len(age_distribution))
        for t in range(len(years)):
            year = start_year + t
            fleet_composition = np.roll(fleet_composition, 1)
            if transition_start <= year <= transition_end:
                fleet_composition[0] = min(100, start_ev_percentage + alpha * (year - start_year) +
                                           beta * (year - start_year) ** 1.2 +
                                           (target_fleet_ev_percentage - start_ev_percentage) * ((year - transition_start) / (transition_end - transition_start))**2) / 100
            else:
                fleet_composition[0] = min(fleet_composition[0], 100)
        return np.dot(fleet_composition, age_distribution) * 100

    def objective_function(params):
        alpha, beta, transition_start, transition_end = params
        return abs(compute_required_growth(alpha, beta, transition_start, transition_end) - target_fleet_ev_percentage)

    bounds = [(0, 5), (0, 5), (start_year, target_year - 10), (start_year + 5, target_year)]
    result = opt.minimize(objective_function, [0.5, 0.1, start_year + 5, target_year - 5], bounds=bounds, method='L-BFGS-B')
    best_alpha, best_beta, best_transition_start, best_transition_end = result.x

    ev_market_shares = []
    for year in years:
        if best_transition_start <= year <= best_transition_end:
            share = min(100, start_ev_percentage + best_alpha * (year - start_year) +
                        best_beta * (year - start_year) ** 1.2 +
                        (target_fleet_ev_percentage - start_ev_percentage) * ((year - best_transition_start) / (best_transition_end - best_transition_start))**2)
        else:
            share = min(ev_market_shares[-1], 100) if ev_market_shares else start_ev_percentage
        ev_market_shares.append(share)

    fleet_composition_final = np.zeros(len(age_distribution))
    cumulative_fleet_ev_percentages = []
    for t in range(len(years)):
        fleet_composition_final = np.roll(fleet_composition_final, 1)
        fleet_composition_final[0] = ev_market_shares[t] / 100
        cumulative_fleet_ev = np.dot(fleet_composition_final, age_distribution) * 100
        cumulative_fleet_ev_percentages.append(cumulative_fleet_ev)

    cumulative_fleet_ev_percentages[-1] = target_fleet_ev_percentage
    return pd.DataFrame({"Year": years, "EV_Market_Share": ev_market_shares, "Cumulative_Fleet_EV_Percentage": cumulative_fleet_ev_percentages})

def update_avft_with_ev_growth(target_fleet_ev_percentage, start_year, target_year, moves_input_dir, moves_excel_spreadsheet):
    """
    Updates the AVFT file by:
    - Computing EV market share growth using a smooth transition function.
    - Adjusting fuel type shares proportionally to maintain sum = 1.
    - Replacing only fuelEngFraction for sourceTypeID = 21, keeping others unchanged.

    Parameters:
    - target_fleet_ev_percentage: The desired fleet EV share in the target year.
    - target_year: The year when the target EV share should be reached.
    - start_year: The year from which the transition begins.
    - age_distribution_file: Path to the age distribution file (CSV).
    - avft_file: Path to the AVFT Excel file.
    """

    # Load the vehicle age distribution
    age_distribution_df = pd.read_csv(os.path.join(moves_input_dir,  "ageDistribution_" + str(target_year) + ".csv"), header=0)
    age_distribution_df["sourceTypeID"] = pd.to_numeric(age_distribution_df["sourceTypeID"], errors="coerce")
    age_distribution_df["yearID"] = pd.to_numeric(age_distribution_df["yearID"], errors="coerce")
    age_distribution_df["ageFraction"] = pd.to_numeric(age_distribution_df["ageFraction"], errors="coerce")

    # Extract age distribution for passenger cars (sourceTypeID = 21) in target_year
    age_distribution = age_distribution_df[
        (age_distribution_df["sourceTypeID"] == 21) & (age_distribution_df["yearID"] == target_year)
    ]["ageFraction"].values

    # Load full AVFT data
    avft_df = pd.read_excel(os.path.join(moves_input_dir,  "avft_" + str(target_year) + ".xlsx"), sheet_name="AVFT")

    # Extract the actual starting EV share for start_year
    ev_fuel_type_id = 9
    ev_share_start = avft_df.loc[
        (avft_df["modelYearID"] == start_year) & (avft_df["fuelTypeID"] == ev_fuel_type_id),
        "fuelEngFraction"
    ].values
    start_ev_percentage = ev_share_start[0] * 100 if len(ev_share_start) > 0 else 5

    # Extract only passenger cars (sourceTypeID = 21) to modify
    avft_passenger_cars = avft_df[avft_df["sourceTypeID"] == 21].copy()

    # Run EV market share function
    ev_market_share_results = linear_ev_market_share_smoothed(
        target_fleet_ev_percentage, start_year, target_year, start_ev_percentage, age_distribution
    )

    # Update only `fuelEngFraction` for passenger cars (sourceTypeID = 21)
    ev_market_share_dict = dict(zip(ev_market_share_results["Year"], ev_market_share_results["EV_Market_Share"]))
    for year in range(start_year, target_year + 1):
        year_mask = avft_passenger_cars["modelYearID"] == year
        ev_share = ev_market_share_dict.get(year, None)
        if ev_share is not None:
            avft_passenger_cars.loc[(year_mask) & (avft_passenger_cars["fuelTypeID"] == ev_fuel_type_id), "fuelEngFraction"] = ev_share / 100
            remaining_fraction = 1 - (ev_share / 100)
            non_ev_mask = (year_mask) & (avft_passenger_cars["fuelTypeID"] != ev_fuel_type_id)
            non_ev_total_before = avft_passenger_cars.loc[non_ev_mask, "fuelEngFraction"].sum()
            if non_ev_total_before > 0:
                avft_passenger_cars.loc[non_ev_mask, "fuelEngFraction"] *= remaining_fraction / non_ev_total_before

    # Merge the updated passenger car data back into the full AVFT dataset
    avft_df.update(avft_passenger_cars)

    # Save the full AVFT sheet with all vehicle types
    wb = load_workbook(moves_excel_spreadsheet)
    if "avft" in wb.sheetnames:
        del wb["avft"]
    wb.save(moves_excel_spreadsheet)

    with pd.ExcelWriter(moves_excel_spreadsheet, engine="openpyxl", mode="a") as writer:
        avft_df.to_excel(writer, sheet_name="avft", index=False)


if __name__ == "__main__":

    args = sys.argv
    print(args)
    
    ev_share = int(args[1])  
    start_year = int(args[2])
    scenario_year = int(args[3])
    moves_input_path = args[4]    
    moves_excel_spreadsheet = args[5]
    
    update_avft_with_ev_growth(ev_share, start_year, scenario_year, moves_input_path, moves_excel_spreadsheet)
